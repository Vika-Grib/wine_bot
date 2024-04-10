# -*- coding: utf-8 -*-

import sqlite3
import time
import requests
from aiogram.types import KeyboardButton, ReplyKeyboardMarkup, ReplyKeyboardRemove
from config import token, api_chat_gpt
import logging
import openpyxl
import parser_BeautifulSoup as pB
import openai
from aiogram import Bot, Dispatcher, types
from aiogram.contrib.fsm_storage.memory import MemoryStorage
from aiogram.dispatcher import FSMContext
from aiogram.dispatcher.filters.state import State, StatesGroup


# Initialize bot and dispatcher
bot = Bot(token=token)
memory_storage = MemoryStorage()
dp = Dispatcher(bot, storage=memory_storage)


# Установка API-ключа OpenAI
openai.api_key = api_chat_gpt


# Дополнительное состояние для викторины
class QuizState(StatesGroup):
    question = State()

# Вопросы викторины
quiz_questions = [
    {"question": "Какой регион в Италии известен своими винами Бароло?", "answer": "Пьемонт"},
    {"question": "Из какого винограда традиционно делают Шампанское?", "answer": "Шардоне"},
    {"question": "Какое вино считается самым старым в мире?", "answer": "Коммандария"},
    #  еще вопросы
]

@dp.message_handler(text_contains='Quiz')
async def quiz_start(message: types.Message, state: FSMContext):
    # Приглашение к викторине
    welcome_message = "Добро пожаловать в винную викторину! 🍷\nГотовы проверить свои знания о вине? Отвечайте на вопросы и узнайте, насколько хорошо вы знаете вино. Вперед к первому вопросу!"
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True, one_time_keyboard=True)
    play_button = types.KeyboardButton("Давай поиграем!")
    not_now_button = types.KeyboardButton("В другой раз")
    markup.add(play_button, not_now_button)
    markup.add(exit_button)
    await message.answer(welcome_message, reply_markup=markup)
    await state.reset_state()

@dp.message_handler(lambda message: message.text == "Давай поиграем!")
async def quiz_play(message: types.Message):
    await QuizState.question.set()
    await message.answer(quiz_questions[0]['question'], reply_markup=exit_markup)

@dp.message_handler(lambda message: message.text == "В другой раз")
async def quiz_not_now(message: types.Message):
    # Пользователь выбрал не играть сейчас, возвращаем в главное меню
    markup = types.ReplyKeyboardMarkup(resize_keyboard=True)
    # Добавьте кнопки главного меню
    btn1 = types.KeyboardButton("Get Wine Rating")
    btn2 = types.KeyboardButton("All wine`s ratings from page")
    btn3 = types.KeyboardButton("Ask AI about wine")
    markup.add(btn1, btn2, btn3)
    await message.answer("Хорошо, может быть в другой раз. Нажмите Back to menu чтобы вернутся в главное меню", reply_markup=exit_markup)


@dp.message_handler(state=QuizState.question)
async def quiz_answer(message: types.Message, state: FSMContext):
    if message.text == "Back to menu":
        await state.finish()  # Завершаем состояние викторины
        await start(message, state)  # Возвращаем пользователя в главное меню
        return

    answer = message.text
    async with state.proxy() as data:
        # Проверяем, есть ли уже заданный вопрос в контексте
        if 'current_question' not in data:
            data['current_question'] = 0

        correct_answer = quiz_questions[data['current_question']]['answer']

        if answer.lower() == correct_answer.lower():
            await message.answer("Правильно! 🎉")
        else:
            await message.answer(f"Неправильно 😢 Правильный ответ: {correct_answer}")

        # Переходим к следующему вопросу или завершаем викторину
        data['current_question'] += 1
        if data['current_question'] < len(quiz_questions):
            await message.answer(quiz_questions[data['current_question']]['question'])
        else:
            await message.answer("Викторина окончена! Спасибо за участие.")
            await state.finish()  # Завершаем состояние викторины


# Функция для обработки запросов пользователя с использованием GPT
def get_gpt_response(user_input):
    response = openai.Completion.create(
        engine="gpt-3.5-turbo-instruct",  # Выбрать нужный движок GPT
        # engine="gpt-3.5-turbo-0125",  # Выбрать нужный движок GPT
        prompt=user_input,
        max_tokens=500,
        temperature=0.7
    )
    return response.choices[0].text.strip()



def get_wine_names_from_api_vivino(url): # запрос работает быстрее парсера
    response = requests.get(url)
    current_page_wines = response.json()['explore_vintage']['matches']
    wine_list = []

    for wine in current_page_wines:
        print(wine)
        name = wine['vintage']['seo_name'].replace('-', ' ')
        rating = wine['vintage']['statistics']['ratings_average']
        wine_info = f"{name} - Rating: {rating}"
        wine_list.append(wine_info)

    for wine_info in wine_list:
        print(wine_info)

    return wine_list

    #print(response.json()['explore_vintage']['matches'][1])
#get_wine_names_from_api_vivino('https://www.vivino.com/webapi/explore/explore?country_code=BY&currency_code=BYN&grape_filter=varietal&min_rating=1&order_by=ratings_average&order=desc&price_range_max=1000&price_range_min=0&wine_type_ids%5B%5D=1&wine_type_ids%5B%5D=2&wine_type_ids%5B%5D=3&wine_type_ids%5B%5D=4&wine_type_ids%5B%5D=24&wine_type_ids%5B%5D=7&page=1&language=en')



# Функция для обработки URL-адреса винной страницы
def excel_paste(page_num, sheet):
    row_index = 2  # Начнем с 2 строки (после заголовков) (чтобы какждый раз не обнулялась - выносим вне функции и делаем в функции глобальной, чтобы она каждый раз не сбрасывалась, а олин раз начала с 2 и дальше ув +1)
    list_of_wines = []
    # global row_index, list_of_wines
   # url = f'https://www.vivino.com/webapi/explore/explore?country_code=PL&currency_code=PLN&grape_filter=varietal&min_rating=1&order_by=ratings_average&order=desc&price_range_max=120&price_range_min=30&wine_type_ids%5B%5D=1&wine_type_ids%5B%5D=2&page={page_num}&language=en'
    url = f'https://www.vivino.com/webapi/explore/explore?country_code=PL&currency_code=PLN&grape_filter=varietal&order_by=price&order=asc&wine_type_ids%5B%5D=1&wine_type_ids%5B%5D=2&wine_type_ids%5B%5D=3&wine_type_ids%5B%5D=4&wine_type_ids%5B%5D=5&wine_type_ids%5B%5D=6&wine_type_ids%5B%5D=7&wine_type_ids%5B%5D=&wine_type_ids%5B%5D=24&page={page_num}&language=en'
    wine_names_and_ratings = get_wine_names_from_api_vivino(url)  # Получаем имена и рейтинги вин с использованием API

    if wine_names_and_ratings:
        for wine_info in wine_names_and_ratings:
            wine_name, rating = wine_info.split(' - Rating: ') # разделяем информацию о вине на имя и рейтинг с помощью .split(' - Rating: '), а затем сохраняем их в файл Excel
            if wine_name in list_of_wines:
                continue

            sheet.cell(row=row_index, column=1, value=wine_name)
            sheet.cell(row=row_index, column=2, value=rating)
            row_index += 1
            # print('new wine: ', wine_name)


def parser():
    list_of_wines = []

    file_name = 'wine_data.xlsx'
    start_page = 1
    last_page = 81
    # Excel и активный лист
    workbook = openpyxl.load_workbook(file_name)  # нужно чтобы открылась таблица ДО парсера, а закрывалась уже только после парсера всех страниц
    sheet = workbook.active

    # Заголовки
    sheet['A1'] = 'Wine Name'
    sheet['B1'] = 'Rating'

    column_A = sheet['A']
    for elem in range(len(column_A)):
        list_of_wines.append(column_A[elem].value) # value - значение ячейки, а не сама ячейка

    row_index = len(list_of_wines) + 1  # определяет последнее значение - то есть записываются вина и после 2026
                                        # - следуующие будут дозаписываться уже начиная с 2027

    for page in range(start_page, last_page+1):
        excel_paste(page, sheet)

    # Сохраняем файл Excel с данными о винах
    workbook.save(file_name)
#parser()


# Кнопка "Exit"
exit_button = KeyboardButton("Back to menu")
exit_markup = ReplyKeyboardMarkup(resize_keyboard=True).add(exit_button)

class ConversationState(StatesGroup):
    conversation = State()


conn = sqlite3.connect('wine_database.db')
# Создаем объект cursor, который позволяет нам взаимодействовать с базой данных и добавлять записи
cursor = conn.cursor()


# команда старт
@dp.message_handler(commands=['start'])
@dp.message_handler(text_contains='Back to menu')   # чтобы просто нажатие на кнопку бэк ту м  срабатывало всегда, даже ничего не заканчивая
@dp.message_handler(text_contains='Back to menu', state='prepared_rating')
@dp.message_handler(text_contains='Back to menu', state='prepared_rating_list')
@dp.message_handler(text_contains='Back to menu', state='conversation')
@dp.message_handler(text_contains='Back to menu', state='quiz')
@dp.message_handler(text_contains='Back to menu', state=ConversationState.conversation.state)
async def start(message: types.Message, state: FSMContext):
    await state.finish()
    try:
        cursor.execute(f'SELECT * FROM users WHERE user_id={message.from_user.id}') # чтобы не пзаписывало много раз одного чела
        user = list(cursor.fetchall())[0]
    except Exception as e:
        print(e)
        cursor.execute(f'''INSERT INTO users VALUES ({str(message.from_user.id)},'{str(message.from_user.username)}','','')''')
        conn.commit()
    user = message.from_user
    markup = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
    btn1 = types.KeyboardButton("Get Wine Rating")
    btn2 = types.KeyboardButton("All wine`s ratings from page")
    btn3 = types.KeyboardButton("Ask AI about wine")
    btn4 = types.KeyboardButton("Quiz")
    markup.row(btn1, btn2)
    markup.add(btn3, btn4)
    await bot.send_message(chat_id=message.from_user.id, text=f"Hello, {message.from_user.first_name}!\nI`m bot. I`m here to help you to find out wine rating and answer questions about wine.", reply_markup=markup)
    # Please send me the URL of a wine page.



@dp.message_handler(commands=['admin'])
@dp.callback_query_handler(text_contains="admin_menu")
@dp.callback_query_handler(text_contains="admin_menu", state='wait_for_photo')
async def admin(message: types.Message, state: FSMContext):
    await state.finish()
    a = types.ReplyKeyboardRemove()
    await bot.send_message(chat_id=message.from_user.id, text='Убираем ненужные кнопки...', reply_markup=a)
    time.sleep(1) # на 1 сек он засыпает / зависнет
    cursor.execute('''SELECT * FROM users''')  # мы подключаемся к нужной таблице - выбираем всех юзеров - получаем список юзеров со всеми параметрами
    all_users = cursor.fetchall()
    users_count = len(all_users)
    #print(all_users)
    markup = types.InlineKeyboardMarkup(row_width=2, resize_keyboard=True)
    btn1 = types.InlineKeyboardButton(text="Список пользователей", callback_data="user_list")
    btn2 = types.InlineKeyboardButton(text="Рассылка всем", callback_data="broadcast_all")
    markup.add(btn1, btn2)
    await bot.send_message(chat_id=message.from_user.id, text=f'''Приветствуем в АДМИНКЕ
Количество пользователей: {users_count}''', reply_markup=markup)


@dp.callback_query_handler(text_contains= "user_list")
async def user_list(callback_query: types.CallbackQuery, state: FSMContext):
    cursor.execute('''SELECT * FROM users''')  # мы подключаемся к нужной таблице - выбираем всех юзеров - получаем список юзеров со всеми параметрами
    all_users = cursor.fetchall()
    answer = ""  # создаем переменную для ответа
    for user in all_users:
        user_id = user[0]
        user_name = user[1]
        answer += f'''{user_id} | @{user_name}\n'''
    markup = types.InlineKeyboardMarkup(row_width=2, resize_keyboard=True)
    btn1 = types.InlineKeyboardButton(text="Вернуться в меню админки", callback_data="admin_menu")
    markup.add(btn1)
    await bot.send_message(chat_id=callback_query.from_user.id, text=answer, reply_markup=markup)


# Command handler to start broadcasting custom information
@dp.callback_query_handler(text_contains="broadcast_all")
async def cmd_broadcast(callback_query: types.CallbackQuery, state: FSMContext):
    keyboard = types.InlineKeyboardMarkup(row_width=2)
    # 👋🖥🔐💼🌐📝💳💵✅📩🔗💰⚖️💻🤵⚙️🌇⏰⚠️📥 📤 💌 ◀️✏️📌🖊 ❗️ ❌🔙
    pay_button = types.InlineKeyboardButton(text="❌ Отмена рассылки", callback_data="admin_menu")
    keyboard.add(pay_button)
    await bot.send_message(chat_id=callback_query.from_user.id, text=f'''
‼️ Бот в ожидании рассылки!
🖼 Пришлите фото для рассылки (Не более 10 фото) !
📄 После фото отправьте текст вашей рассылки.''', reply_markup=keyboard)
    await state.set_state('wait_for_photo')


@dp.message_handler(content_types=['photo'], state="wait_for_photo")
async def photo_handler(message: types.Message, state: FSMContext):
    # we are here if the first message.content_type == 'photo'
    # save the largest photo (message.photo[-1]) in FSM, and start photo_counter
    await state.update_data(photo_0=message.photo[-1].file_id, photo_counter=0, text='')
#     await bot.send_message(chat_id=message.from_user.id, text=f'''
# 🖼 Пришлите следующее фото или текст рассылки !
# ‼️После отправки текста и ссылки ваша рассылка будет автоматически разослана.''')
    await state.set_state('next_photo')

@dp.message_handler(content_types=['photo'], state='next_photo')
async def next_photo_handler(message: types.Message, state: FSMContext):
    # we are here if the second and next messages are photos
    async with state.proxy() as data:
        data['photo_counter'] += 1
        if data['photo_counter'] == 1:
            await bot.send_message(chat_id=message.from_user.id, text=f'''
🖼 Пришлите следующее фото или текст рассылки !
‼️После отправки текста и ссылки ваша рассылка будет автоматически разослана.''')
        photo_counter = data['photo_counter']
        data[f'photo_{photo_counter}'] = message.photo[-1].file_id

    await state.set_state('next_photo')

# 👋🖥🔐💼🌐📝💳💵✅📩🔗💰⚖️💻🤵
@dp.message_handler(content_types=["text"], state='next_photo')
async def not_foto_handler(message: types.Message, state: FSMContext):
    # we are here if the second and next messages are not photos
    ad_text = message.text
    async with state.proxy() as data:
        data['text'] = ad_text

        await bot.send_message(chat_id=message.from_user.id, text='''
Теперь введите ссылку:
("-" если не хотите вводить)''')
        await state.set_state("url")


@dp.message_handler(content_types=["text"], state='url')
async def not_foto_handler(message: types.Message, state: FSMContext):
    # we are here if the second and next messages are not photos
    url = message.text

    async with state.proxy() as data:
        ad_text = data['text']
        # here we do something with data dictionary with all photos
        keyboard = types.InlineKeyboardMarkup(row_width=2)
        pay_button4 = types.InlineKeyboardButton(text="🌐 Меню админа 🌐", callback_data="admin_menu")
        keyboard.add(pay_button4)
        await bot.send_message(chat_id=message.from_user.id, text='✅ Ваша рассылка разослана!', reply_markup=keyboard)
        await broadcast_custom_information(ad_text, data, url)
        await state.finish()


async def broadcast_custom_information(message_to_broadcast, photos_ids, url):
    all_users_ids = []
    try:
        # Retrieve a list of users who interacted with the bot
        cursor.execute(f'''Select * from users ''')
        all_users = list(cursor.fetchall())

        for user in all_users:
            if user[0] != 0:  # проверка на созданную вручную запись с id
                all_users_ids.append(user[0])

    except Exception as e:
        logging.exception(f"Failed to get all users: {e}")  # на сервере можно было видеть логи - при запуске файла он будет создавать файлик .log, чтобы видеть любые логи, выводы

    logging.info(f"Total {len(all_users_ids)} users found")

    media = types.MediaGroup() # сюда добавляются все фото
    photos_ids = list(photos_ids.values())[:1] + list(photos_ids.values())[3:]
    if url == "-":
        for i in range(len(photos_ids)):
            # print(photos_ids[i])
            if i == 0:
                media.attach_photo(types.InputMediaPhoto(photos_ids[i], caption=message_to_broadcast))  # attach_photo - когда id = 0, т.е. добавляем первое фото, мы доб подпись нашим текстом
            else:
                media.attach_photo(types.InputMediaPhoto(photos_ids[i]))
        # Send the custom information message to all users
        for user_id in all_users_ids:
            try:
                await bot.send_media_group(chat_id=user_id, media=media)  # метод для отправки альбома фотографий
            except Exception as e:
                print(e)
                if "bot was blocked by the user" in str(e):
                    cursor.execute(f'''UPDATE users SET blocked = "Yes" WHERE user_Id={user_id}''')
                    logging.exception(f"Failed to send message to user {user_id}: {e}")
    else:
        for i in range(len(photos_ids)):
            if i == 0:
                media.attach_photo(types.InputMediaPhoto(photos_ids[i]))
            else:
                media.attach_photo(types.InputMediaPhoto(photos_ids[i]))
        # Send the custom information message to all users
        for user_id in all_users_ids:
            try:
                await bot.send_media_group(chat_id=user_id, media=media)
                # await bot.send_media_group(chat_id=user_id, media=params)
            except Exception as e:
                print(e)
                if "bot was blocked by the user" in str(e):
                    cursor.execute(f'''UPDATE users SET blocked = "Yes" WHERE user_Id={user_id}''')
                    logging.exception(f"Failed to send message to user {user_id}: {e}")

                logging.exception(f"Failed to send message to user {user_id}: {e}")
            keyboard = types.InlineKeyboardMarkup(row_width=2)
            pay_button1 = types.InlineKeyboardButton(text="ПЕРЕЙТИ", url=str(url))
            keyboard.add(pay_button1)
            await bot.send_message(chat_id=user_id, text=message_to_broadcast, reply_markup=keyboard)




# Обработчик нажатия кнопки "Ask AI about wine"
@dp.message_handler(text_contains='Ask AI about wine')
async def process_message(message: types.Message, state: FSMContext):
    await message.answer('Welcome! Feel free to ask me anything about wine.', reply_markup=exit_markup)
    await ConversationState.conversation.set()


# Обработчик состояния "conversation"
@dp.message_handler(content_types=['text'], state=ConversationState.conversation)
async def process_message(message: types.Message, state: FSMContext):
    # Проверяем наличие кнопки "Exit"

    cursor.execute(f'''Select * from users WHERE user_id={message.from_user.id}''')
    user = list(cursor.fetchall())[0]
    theme = user[2]
    main_idea = user[3]
    if theme == '' and main_idea == '':
        first_gpt_response = get_gpt_response(message.text)
        print(first_gpt_response, '1_GPT_RESPONSE!!!!')
        await message.answer(first_gpt_response.strip(), reply_markup=exit_markup)
        # Сохраняем ответ gpt для будущих запросов
#
#             gpt_response = get_gpt_response(f'''
# Используя методы анализа, демонстрируемые в примерах ниже, определите основную тему и сформулируйте главную мысль кратко и ясно из пользовательского сообщения.
#
# Пользовательское сообщение: [{first_gpt_response}]
# Твоя задача:
# 1. Определите основную тему данного текста и выделите ее в явном виде.
# 2. Описать основную мысль текста кратко и ясно.
#
# Ожидаемый формат ответа:
# Тема: Тема текста
# Основная мысль: Краткое описание основной мысли текста
#
# Пример идеальных ответов:
# Пример 1:
# Текст: Я бы посоветовал попробовать красные вина из сорта Темпранильо, которые производятся в регионе Риоха. Они являются одними из наиболее известных и высококачественных в Испании. Также можно попробовать белые и розовые вина из уникальных сортов винограда, которые также производятся в Риохе. Хорошим выбором будет и вино из региона Рибера-дель-Дуэро, также известное своими красными винами из Темпранильо. В общем, я бы посоветовал попробовать различные сорта и стили вин из разных регионов Испании, чтобы познакомиться с их разнообразием и насладиться уникальным вкусом каждого из них.
# Тема: Испанские вина.
# Основная мысль: Рекомендуется исследовать разнообразие испанских вин, начиная с Темпранильо из Риохи и Рибера-дель-Дуэро.
#
# Пример 2:
# Текст: Начать пробовать красные супервина можно с французских вин, таких как Шато Латур или Шато Марго, итальянских вин, например, Бароло или Брунелло ди Монтальчино, аргентинских вин, например, Малбек или Каберне Совиньон, и других регионов, известных своими высококачественными красными винами, например, Напа Вэлли в Калифорнии или Бордо во Франции.
# Тема: красные супервина.
# Основная мысль: Знакомство с красными винами можно начать с Шато Латур, Шато Марго, Бароло, Брунелло ди Монтальчино, Мальбек, Каберне Совиньон, Напа Вэлли - Калифорния, Бордо - Франция.
#
# ''')

        gpt_response = get_gpt_response(f'''
Выдели тему сообщения в квадратных скобках, затем кратко запиши основную мысль.
Пример:
Начать пробовать красные супервина можно с французских вин, таких как Шато Латур или Шато Марго, итальянских вин, например, Бароло или Брунелло ди Монтальчино, аргентинских вин, например, Малбек или Каберне Совиньон, и других регионов, известных своими высококачественными красными винами, например, Напа Вэлли в Калифорнии или Бордо во Франции.

Твой идеальный ответ:
Тема: красные супервина.
Основная мысль: Шато Латур, Шато Марго, Бароло, Брунелло ди Монтальчино, Малбек, Каберне Совиньон, Напа Вэлли - Калифорния, Бордо - Франция.

[{first_gpt_response}]
''')
        gpt_theme = gpt_response.split('Основная мысль:')[0].replace('Тема:', '').strip()
        gpt_main_idea = gpt_response.split('Основная мысль:')[1].strip()
        theme = gpt_theme
        main_idea = gpt_main_idea
        cursor.execute(f'''UPDATE users SET theme='{theme}' WHERE user_id = {message.from_user.id}''')
        conn.commit()
        cursor.execute(f'''UPDATE users SET main_idea='{main_idea}' WHERE user_id = {message.from_user.id}''')
        conn.commit()

    else:
        context = f'''
Дай ответ на "Вопрос пользователя". Основная мысль вашего диалога находится в "Контекст".
Контекст: 
Тема: {theme}
Основная мысль: {main_idea}
Вопрос пользователя: 
{message.text}
'''
        # Передаем контекст в GPT
        next_gpt_response = get_gpt_response(context)
        await message.answer(next_gpt_response.strip(), reply_markup=exit_markup)
        gpt_response = get_gpt_response(f'''
Выдели тему сообщения в квадратных скобках, затем кратко запиши основную мысль.
Пример:
Начать пробовать красные супервина можно с французских вин, таких как Шато Латур или Шато Марго, итальянских вин, например, Бароло или Брунелло ди Монтальчино, аргентинских вин, например, Малбек или Каберне Совиньон, и других регионов, известных своими высококачественными красными винами, например, Напа Вэлли в Калифорнии или Бордо во Франции.

Твой идеальный ответ:
Тема: красные супервина.
Основная мысль: Шато Латур, Шато Марго, Бароло, Брунелло ди Монтальчино, Малбек, Каберне Совиньон, Напа Вэлли - Калифорния, Бордо - Франция.

[{next_gpt_response}]
''')
        print(next_gpt_response)
        gpt_theme = gpt_response.split('Основная мысль:')[0].replace('Тема:', '').strip()
        gpt_main_idea = gpt_response.split('Основная мысль:')[1].strip()
        theme += ' + ' + gpt_theme
        main_idea += ' + ' + gpt_main_idea
        cursor.execute(f'''UPDATE users SET theme='{theme}' WHERE user_id = {message.from_user.id}''')
        conn.commit()
        cursor.execute(f'''UPDATE users SET main_idea='{main_idea}' WHERE user_id = {message.from_user.id}''')
        conn.commit()
        print(gpt_response, "!!!!!")

        # Сбрасываем текущее состояние, но продолжаем слушать ввод пользователя
        await state.reset_state()

        # После ответа бота, переводим его снова в состояние "conversation"
    await ConversationState.conversation.set()



@dp.message_handler(text_contains='Get Wine Rating')
async def process_message(message: types.Message, state: FSMContext):
    await bot.send_message(chat_id=message.from_user.id, text='Please send me the name of a wine for rating.', reply_markup=exit_markup)
    await state.set_state('prepared_rating')  # ожидает ссылку и наше состояние устанавливается


@dp.message_handler(content_types=['text'], state='prepared_rating')
async def process_message(message: types.Message, state: FSMContext):
    cursor.execute(f'''Select * from wine_ratings''')
    wine_ratings = list(cursor.fetchall())
    user_wine = message.text # то что вводит юзер - название вина
    print('!!!', user_wine)
    rating = False
    for wine_rating in wine_ratings:
        if user_wine.lower() in wine_rating[0].lower().strip(): # тут мы сравниваем именно имя [0] чтобы в нижнем регистре
            rating = wine_rating[1]
            await bot.send_message(chat_id=message.from_user.id, text=f'Рейтинг {user_wine} - {rating}')
            print(message.from_user.id)
            break
    rating_answer = ''
    if not rating:
        rating_answer = pB.get_vivino_rating(user_wine.lower())  # если в бд нет вина, то мы залазим в нашу функцию поиска р в вивино
        if rating_answer is not None:
            await bot.send_message(chat_id=message.from_user.id, text=f'Рейтинг {user_wine} - {rating_answer}')
            cursor.execute('''INSERT INTO wine_ratings(wine_name, rating) VALUES (?,?)''',
                           (user_wine, rating_answer))
            conn.commit()
    print(rating, rating_answer)
        #print(f"Сравниваем: {wine_rating[0].lower()} с {user_wine.lower()}")
    await state.finish()


@dp.message_handler(text_contains='All wine`s ratings from page')
async def process_message(message: types.Message, state: FSMContext):
    await bot.send_message(chat_id=message.from_user.id, text='Please send me the URL of a wine page.', reply_markup=exit_markup)
    await state.set_state('prepared_rating_list')  # ожидает ссылку и наше состояние устанавливается


@dp.message_handler(content_types=['text'], state='prepared_rating_list')
async def process_message(message: types.Message, state: FSMContext):
    user_url = message.text
    if user_url.startswith('http'):
        # чтобы запомнить предыдущий ответ
        await state.update_data(user_url=user_url)
        await bot.send_message(chat_id=message.from_user.id, text='Great! Now, please copy as it`s named on page and send it to me.')
        await state.set_state('completed_request')  # ожидает ссылку и наше состояние устанавливается
    else:
        await bot.send_message(chat_id=message.from_user.id, text="Please send valid URL.")


@dp.message_handler(content_types=['text'], state='completed_request')
async def process_message(message: types.Message, state: FSMContext):
    sended_message = await bot.send_message(chat_id=message.from_user.id, text='Please, wait. Your wine is coming..')
    sended_animation = await bot.send_animation(chat_id=message.from_user.id, animation='https://i.pinimg.com/originals/c7/4d/91/c74d9176c3c98ca4724ddebfe403192e.gif')
    async with state.proxy() as data:  # при помощи state.proxy() as data достали все аргументы из текущего состояния и записали в data
        user_url = data["user_url"]  # достаем из data нашу информацию
    if user_url.startswith('http'):
        wine_class_name = pB.get_wine_class_name(message, user_url)
        # print("wcm: ", wine_class_name)
        wine_page_list = pB.collect_wine_names(user_url, wine_class_name)
        rating_page_list = []
        for wine in wine_page_list:
            rating_page_list.append(pB.get_vivino_rating(wine))
        response = pB.generate_response(wine_page_list, rating_page_list)

        await bot.edit_message_text(chat_id=message.from_user.id, text=response, message_id=sended_message.message_id)  # тут мы редактируем уже отправленное наше сообщение text='Please, wait. Your wine is coming..'
        await bot.delete_message(chat_id=message.from_user.id, message_id=sended_animation.message_id)
        await state.finish()  # заканчиваем стейт, чтобы он сбросился
    else:
        await bot.send_message(chat_id=message.from_user.id, text="Please send valid URL.")
        await state.set_state('prepared_rating_list')


# Run the bot
if __name__ == '__main__':
    from aiogram import executor
    # print(get_gpt_response(input()))
    # настройка журнала
    logging.basicConfig(format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
                        level=logging.INFO)
    # logging.info('Starting the bot...')
    executor.start_polling(dp, skip_updates=True)



